from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
import re

TWEET_XPATHS = {
    "tweet_article": '//article[@data-testid="tweet"]',
    "user_name": './/div[@data-testid="User-Name"]',
    "tweet_text": './/div[@data-testid="tweetText"]',
    "tweet_time": './/time',
    "tweet_time_parent": './/time/parent::a'
}


def limpiar_texto(texto: str) -> str:
    """
    Limpia caracteres invisibles comunes en textos de tweets.
    """
    return texto.replace("\u2066", "").replace("\u2069", "").strip()


def es_retweet(tweet_element: WebElement) -> bool:
    """
    Determina si un tweet es un retweet o repost.
    """
    try:
        social_context = tweet_element.find_element(By.XPATH, './/div[@data-testid="socialContext"]')
        texto = social_context.text.strip().lower()
        return any(palabra in texto for palabra in ["reposte", "retwitte", "retweeted", "reposteó", "retwitteó"])
    except:
        return False


def es_tweet_fijado(tweet_element: WebElement, timeout: int = 2) -> bool:
    """
    Verifica si un tweet está fijado (pinned).
    """
    try:
        social_context = WebDriverWait(tweet_element, timeout).until(
            EC.presence_of_element_located((By.XPATH, './/div[@data-testid="socialContext"]'))
        )
        texto = social_context.text.strip().lower()
        return "fijado" in texto or "pinned" in texto
    except:
        return False


def hacer_scroll(driver, repeticiones=5, pausa=1.0):
    """
    Realiza múltiples desplazamientos hacia abajo en la página.
    """
    body = driver.find_element(By.TAG_NAME, 'body')
    for _ in range(repeticiones):
        body.send_keys(Keys.PAGE_DOWN)
        time.sleep(pausa)


def guardar_tweets_excel(tweets_data, temp_file):
    if os.path.exists(temp_file):
        book = load_workbook(temp_file)
        sheet = book["Posts"]
    else:
        book = Workbook()
        sheet = book.active
        sheet.title = "Posts"
        sheet.append([
            "ID_Político", "Enlace_Post", "Contenido", "Fecha_Publicación",
            "Comentarios_Extraidos", "Comentarios_Totales", "Retweets", "Likes"
        ])

    for tweet in tweets_data:
        fila = tweet[:4] + [0] + tweet[4:]
        sheet.append(fila)

    book.save(temp_file)
    book.close()


def fusionar_archivos(file_path):
    """Fusiona todos los archivos temporales en el archivo principal"""
    archivos_temporales = [f for f in os.listdir() if f.startswith("politicos_") and f.endswith(".xlsx")]

    # Cargar el archivo principal si existe
    if os.path.exists(file_path):
        df_principal = pd.read_excel(file_path, sheet_name="Posts")
    else:
        df_principal = pd.DataFrame(columns=["ID_Político", "Enlace_Post", "Contenido", "Fecha_Publicación"])

    # Cargar y combinar archivos temporales
    for temp_file in archivos_temporales:
        df_temp = pd.read_excel(temp_file, sheet_name="Posts")
        df_principal = pd.concat([df_principal, df_temp], ignore_index=True)

        # Eliminar archivo temporal después de fusionarlo
        os.remove(temp_file)

    # Guardar el resultado final
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_principal.to_excel(writer, sheet_name="Posts", index=False)

    print(f"✅ Fusión completada. Datos guardados en {file_path}")


def dividir_cuentas_entre_procesos(cuentas, num_procesos):
    """Distribuye las cuentas en bloques consecutivos lo más equilibrado posible entre los procesos."""
    cuentas_ordenadas = sorted(cuentas, key=lambda x: x["ID_Político"])
    total = len(cuentas_ordenadas)

    base = total // num_procesos
    extra = total % num_procesos

    resultado = []
    inicio = 0

    for i in range(num_procesos):
        fin = inicio + base + (1 if i < extra else 0)
        resultado.append(cuentas_ordenadas[inicio:fin])
        inicio = fin

    return resultado


def es_solo_enlace(texto):
    """Determina si el texto es únicamente un enlace."""
    texto = texto.strip()
    return bool(re.fullmatch(r'https?://\S+', texto))
