import time
import random
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from extractor.posts.ApoyoPosts import limpiar_texto, es_retweet, es_tweet_fijado, hacer_scroll, TWEET_XPATHS, \
    guardar_tweets_excel, fusionar_archivos, es_solo_enlace
from extractor.FuncionesConjuntas import get_chrome_driver, cargar_cookies, extraer_metricas_tweet


class TweetExtractor:
    def __init__(self, file_path, driver_path, cookies, process_id=0, headless=True, cantidad=16):
        self.file_path = file_path
        self.driver_path = driver_path
        self.cookies = cookies
        self.process_id = process_id
        self.cantidad = cantidad
        self.driver = get_chrome_driver(process_id, driver_path=driver_path, headless=headless)
        self.temp_file = f"politicos_{process_id}.xlsx" if headless else "politicos_individual.xlsx"
        self.XPATHS = TWEET_XPATHS

    def _cargar_id_politico(self, username):
        df_meta = pd.read_excel(self.file_path, sheet_name="Metadata")
        df_meta["ID_Político"] = df_meta["ID_Político"].ffill().astype(int)
        match = df_meta[df_meta["Twitter"].str.lower() == username.lower()]
        if match.empty:
            raise ValueError(f"❌ No se encontró el username '{username}' en Metadata.")
        return int(match["ID_Político"].iloc[0])

    def _cargar_tweets_existentes(self, politicos_ids):
        existing = set()
        if os.path.exists(self.file_path):
            try:
                df = pd.read_excel(self.file_path, sheet_name="Posts")
                df_filtrado = df[df["ID_Político"].isin(politicos_ids)]
                existing.update(df_filtrado["Enlace_Post"].dropna().tolist())
            except:
                pass
        if os.path.exists(self.temp_file):
            book = load_workbook(self.temp_file)
            sheet = book["Posts"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                existing.add(row[1])
        return existing

    def _extraer_para(self, username, id_politico, existing_links, cantidad):
        driver = self.driver

        self.inicializar_driver_y_cookies(driver)

        self.ir_a_perfil(driver, username)

        tweet_count = 0
        scroll_attempts = 0
        max_scroll_attempts = 100
        tweets_data = []

        while tweet_count < cantidad and scroll_attempts < max_scroll_attempts:
            try:
                tweets_container = self.obtener_tweets_visibles(driver)
            except:
                scroll_attempts += 1
                print(f"🔄 @{username} – Scroll #{scroll_attempts} (no se encontraron tweets)")
                hacer_scroll(driver, 5, 1)
                continue

            nuevos_tweets = 0

            nuevos_tweets, scroll_attempts, tweet_count = self.procesar_lote_de_tweets(cantidad, existing_links,
                                                                                       id_politico, nuevos_tweets,
                                                                                       scroll_attempts, tweet_count,
                                                                                       tweets_container, tweets_data,
                                                                                       username)

            if nuevos_tweets == 0:
                scroll_attempts += 1
                if scroll_attempts >= max_scroll_attempts:
                    print(f"⚠️ @{username} – Límite de scroll alcanzado ({max_scroll_attempts})")
                    break
                print(f"🔄 @{username} – Scroll adicional #{scroll_attempts}")
                hacer_scroll(driver, 5, 1)

        print(f"✅ Extracción finalizada para @{username} - Total descargados: {tweet_count}")
        return tweets_data

    def procesar_lote_de_tweets(self, cantidad, existing_links, id_politico, nuevos_tweets, scroll_attempts,
                                tweet_count, tweets_container, tweets_data, username):
        for tweet in tweets_container:
            if tweet_count >= cantidad:
                break
            try:
                if es_retweet(tweet):
                    print(f"🔁 @{username} – Retweet detectado y omitido")
                    continue

                handle = tweet.find_element(By.XPATH, self.XPATHS["user_name"]).text.split("\n")[1]
                if handle.lower() != f"@{username.lower()}":
                    continue

                date_raw = tweet.find_element(By.XPATH, self.XPATHS["tweet_time"]).get_attribute("datetime")
                date = datetime.strptime(date_raw, "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%Y-%m-%d %H:%M:%S")
                tweet_link = tweet.find_element(By.XPATH, self.XPATHS["tweet_time_parent"]).get_attribute("href")
                content = limpiar_texto(tweet.find_element(By.XPATH, self.XPATHS["tweet_text"]).text.strip())

                if tweet_link in existing_links or not content or es_solo_enlace(content):
                    print(f"⚠️ Tweet ya existente o sin contenido de @{username}")
                    continue

                comentarios, retweets, likes = extraer_metricas_tweet(tweet)
                tweets_data.append([id_politico, tweet_link, content, date, comentarios, retweets, likes])
                existing_links.add(tweet_link)
                tweet_count += 1
                nuevos_tweets += 1
                scroll_attempts = 0
                print(f"📥 @{username} {tweet_count}/{cantidad} - {comentarios}/{retweets}/{likes}")
                time.sleep(random.randint(6, 10))
            except:
                continue
        return nuevos_tweets, scroll_attempts, tweet_count

    def ir_a_perfil(self, driver, username):
        print(f"🔍 Comenzando extracción para @{username}")
        driver.get(f"https://x.com/{username}")
        time.sleep(random.randint(8, 12))

    def inicializar_driver_y_cookies(self, driver):
        if not hasattr(self, "_cookies_cargadas"):
            cargar_cookies(driver, self.cookies)
            time.sleep(5)
            self._cookies_cargadas = True

    def extraer_para_usuario(self, username):
        id_politico = self._cargar_id_politico(username)
        tweets_existentes = self._cargar_tweets_existentes([id_politico])
        data = self._extraer_para(username, id_politico, tweets_existentes, self.cantidad)
        print(f"🔍 Iniciando descarga para @{username}")
        guardar_tweets_excel(data, self.temp_file)
        self.driver.quit()
        if data:
            fusionar_archivos(self.file_path)
            print(f"✅ Finalizada la descarga de @{username}")
        else:
            print(f"⚠️ No se extrajo nada de @{username}")

    def extraer_para_grupo(self, politicos):
        politicos_ids = [p["ID_Político"] for p in politicos]
        existing = self._cargar_tweets_existentes(politicos_ids)

        for politico in politicos:
            username = politico["Twitter"]
            id_politico = politico["ID_Político"]
            tweets = self._extraer_para(username, id_politico, existing, self.cantidad)
            guardar_tweets_excel(tweets, self.temp_file)

        self.driver.quit()
        print(f"✅ Finalizado proceso grupal ID {self.process_id}")

    def _extraer_para_entre_fechas(self, username, id_politico, existing_links, fecha_inicio, fecha_fin):
        driver = self.driver
        self.inicializar_driver_y_cookies(driver)

        self.ir_a_perfil(driver, username)

        scroll_attempts = 0
        max_scroll_attempts = 100
        tweets_data = []

        while scroll_attempts < max_scroll_attempts:
            try:
                tweets_container = self.obtener_tweets_visibles(driver)
            except:
                scroll_attempts += 1
                hacer_scroll(driver, 5, 0.5)
                continue

            nuevos_tweets = 0

            for tweet in tweets_container:
                try:
                    handle = tweet.find_element(By.XPATH, self.XPATHS["user_name"]).text.split("\n")[1]
                    if handle.lower() != f"@{username.lower()}":
                        continue

                    date_raw = tweet.find_element(By.XPATH, self.XPATHS["tweet_time"]).get_attribute("datetime")
                    date_obj = datetime.strptime(date_raw, "%Y-%m-%dT%H:%M:%S.%fZ")
                    date = date_obj.strftime("%Y-%m-%d %H:%M:%S")

                    if es_tweet_fijado(tweet):
                        if not (fecha_inicio <= date_obj <= fecha_fin):
                            print(f"📌 Tweet fijado omitido por estar fuera de fechas: {date}")
                            continue

                    if date_obj < fecha_inicio:
                        try:
                            content = limpiar_texto(
                                tweet.find_element(By.XPATH, self.XPATHS["tweet_text"]).text.strip())
                        except:
                            content = "(Sin texto)"

                        print(f"🛑 @{username} – Se alcanzó el límite inferior: {fecha_inicio.date()}")
                        print(f"🗓️ Fecha del tweet: {date}")
                        print(f"✏️ Contenido: {content}")
                        print(
                            f"🔗 Enlace: {tweet.find_element(By.XPATH, self.XPATHS['tweet_time_parent']).get_attribute('href')}")
                        return tweets_data

                    if fecha_inicio <= date_obj <= fecha_fin:
                        tweet_link = tweet.find_element(By.XPATH, self.XPATHS["tweet_time_parent"]).get_attribute(
                            "href")
                        content = limpiar_texto(tweet.find_element(By.XPATH, self.XPATHS["tweet_text"]).text.strip())
                        if tweet_link in existing_links or not content or es_solo_enlace(content):
                            continue

                        comentarios, retweets, likes = extraer_metricas_tweet(tweet)
                        tweets_data.append([id_politico, tweet_link, content, date, comentarios, retweets, likes])
                        existing_links.add(tweet_link)
                        nuevos_tweets += 1
                        print(f"📥 @{username} - {date} - {comentarios}/{retweets}/{likes}")
                        time.sleep(random.randint(3, 5))
                except:
                    continue

            if nuevos_tweets > 0 and scroll_attempts > 0:
                scroll_attempts = 0
                print(f"🔄 @{username} – Nuevos tweets encontrados, reiniciando scrolls")
            else:
                scroll_attempts += 1
                print(f"🔄 Scroll #{scroll_attempts} para @{username} (sin nuevos tweets)")

            if scroll_attempts >= max_scroll_attempts:
                print(f"⚠️ @{username} – Máximo scrolls alcanzado ({max_scroll_attempts})")
                break

            body = driver.find_element(By.TAG_NAME, 'body')
            for _ in range(5):
                body.send_keys(Keys.PAGE_DOWN)
                time.sleep(0.25)

        print(f"✅ Extracción entre fechas finalizada para @{username}")
        return tweets_data

    def obtener_tweets_visibles(self, driver):
        tweets_container = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, self.XPATHS["tweet_article"]))
        )
        return tweets_container

    def extraer_para_usuario_entre_fechas(self, username, fecha_inicio_str, fecha_fin_str):
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d")
        fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d")
        id_politico = self._cargar_id_politico(username)
        tweets_existentes = self._cargar_tweets_existentes([id_politico])
        data = self._extraer_para_entre_fechas(username, id_politico, tweets_existentes, fecha_inicio, fecha_fin)
        guardar_tweets_excel(data, self.temp_file)
        self.driver.quit()
        if data:
            fusionar_archivos(self.file_path)
            print(f"✅ Finalizada la descarga de @{username} entre fechas")
        else:
            print(f"⚠️ No se extrajo nada de @{username}")

    def extraer_para_grupo_entre_fechas(self, politicos, fecha_inicio_str, fecha_fin_str):
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d")
        fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d")
        politicos_ids = [p["ID_Político"] for p in politicos]
        existing = self._cargar_tweets_existentes(politicos_ids)

        for politico in politicos:
            username = politico["Twitter"]
            id_politico = politico["ID_Político"]
            tweets = self._extraer_para_entre_fechas(username, id_politico, existing, fecha_inicio, fecha_fin)
            guardar_tweets_excel(tweets, self.temp_file)

        self.driver.quit()
        print(f"✅ Finalizado proceso grupal entre fechas ID {self.process_id}")
