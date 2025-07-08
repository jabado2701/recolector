import os
import re
import time
import random
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from extractor.FuncionesConjuntas import cargar_cookies, get_chrome_driver, extraer_metricas_tweet
from extractor.comentarios.ApoyoComentarios import es_descubre_mas


def extraer_comentarios(posts_info, file_path, process_id, cookies, driver_path, cantidad_comentarios=10):
    """
    Extrae comentarios de posts de X/Twitter. Detecta respuestas del autor y guarda métricas.
    """
    book, df_posts_dict, driver, existing_comment_links, sheet, start_time, temp_file = puesta_en_marcha(cookies,
                                                                                                         driver_path,
                                                                                                         file_path,
                                                                                                         process_id)

    for post in posts_info:
        post_link = post["Enlace_Post"]
        username_autor = post["Autor_Twitter"]

        comentarios_actuales = df_posts_dict.get(post_link, {}).get("Comentarios_Extraidos", 0)
        if comentarios_actuales >= cantidad_comentarios:
            print(f"⏩ Post omitido, ya tiene {comentarios_actuales} comentarios: {post_link}")
            continue

        print(f"🔗 Procesando post: {post_link}")
        print(f"🔄 Hay {comentarios_actuales}, se extraerán {cantidad_comentarios - comentarios_actuales} más.")
        driver.get(post_link)
        time.sleep(random.randint(4, 6))

        comment_count = comentarios_actuales
        scroll_attempts = 0
        max_scroll_attempts = 3
        first_comment = second_comment = None

        while comment_count < cantidad_comentarios and scroll_attempts < max_scroll_attempts:
            try:
                blocks = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, '//div[@data-testid="cellInnerDiv"]'))
                )
            except Exception:
                print(f"⚠️ No se encontraron bloques en {post_link}.")
                break

            for block in blocks:
                if comment_count >= cantidad_comentarios:
                    break

                if es_descubre_mas(block):
                    print("⛔️ Se alcanzó la sección 'Descubre más'. Fin de comentarios.")
                    break

                try:
                    comment = block.find_element(By.XPATH, './/article[@data-testid="tweet"]')
                except:
                    continue

                try:
                    handle = comment.find_element(By.XPATH, './/div[@data-testid="User-Name"]').text.split("\n")[1]
                    comment_link = comment.find_element(By.XPATH, './/time/parent::a').get_attribute("href")

                    if comment_link in existing_comment_links:
                        continue

                    date_raw = comment.find_element(By.XPATH, './/time').get_attribute("datetime")
                    comment_date = datetime.strptime(date_raw, "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%Y-%m-%d %H:%M:%S")

                    content_els = comment.find_elements(By.XPATH, './/div[@data-testid="tweetText"]')
                    comment_content = content_els[0].text.strip() if content_els else ""

                    if re.fullmatch(r'https?://\S+', comment_content):
                        print("⚠️ Comentario descartado por ser solo un enlace.")
                        continue

                    if not comment_content:
                        if comment.find_elements(By.XPATH, './/div[@data-testid="videoPlayer"]'):
                            print("⚠️ Comentario solo con video, omitido.")
                        else:
                            print("⚠️ Comentario vacío, omitido.")
                        continue

                    comentarios, retweets, likes = extraer_metricas_tweet(comment)

                    current = {
                        "author": handle,
                        "link": comment_link,
                        "content": comment_content,
                        "date": comment_date,
                        "metricas": (comentarios, retweets, likes)
                    }

                    if not first_comment:
                        first_comment = current
                        continue
                    elif not second_comment:
                        second_comment = current
                    else:
                        pass

                    if first_comment["author"].lower() != f"@{username_autor.lower()}":
                        if second_comment["author"].lower() == f"@{username_autor.lower()}":
                            print(f"💬 Respuesta del autor detectada: {second_comment['content']}")
                            respuesta = second_comment["content"]
                            existing_comment_links.update([first_comment["link"], second_comment["link"]])
                        else:
                            respuesta = "No"
                            existing_comment_links.add(first_comment["link"])

                        guardar_comentario(first_comment, post_link, respuesta, sheet)
                        comment_count += 1
                        time.sleep(random.uniform(1.5, 2.5))

                        first_comment = second_comment
                        second_comment = None

                    else:
                        existing_comment_links.add(first_comment["link"])
                        time.sleep(2)
                        first_comment = second_comment
                        second_comment = None

                except Exception as e:
                    print(f"⚠️ Error procesando comentario: {e}")
                    continue

            if comment_count < cantidad_comentarios:
                scroll_attempts += 1
                if scroll_attempts >= max_scroll_attempts:
                    print(f" Límite de scroll alcanzado en {post_link}.")
                    break

                print(f"🔄 Scroll... ({scroll_attempts}/{max_scroll_attempts})")
                body = driver.find_element(By.TAG_NAME, 'body')
                for _ in range(2):
                    body.send_keys(Keys.PAGE_DOWN)
                    time.sleep(0.5)

        comment_count = procesar_comentario_final(cantidad_comentarios, comment_count, existing_comment_links,
                                                  first_comment, post_link, sheet, username_autor)

        print(f"✅ Comentarios procesados: {comment_count}/{cantidad_comentarios} en {post_link}")

    driver.quit()
    book.save(temp_file)
    book.close()
    print(f"⏳ Tiempo total: {time.time() - start_time:.2f} segundos")


def puesta_en_marcha(cookies, driver_path, file_path, process_id):
    """
    Inicializa entorno de trabajo para un proceso de extracción de comentarios.
    Retorna objetos necesarios para continuar la operación.
    """
    start_time = time.time()

    driver = get_chrome_driver(process_id, headless=True, driver_path=driver_path)
    cargar_cookies(driver, cookies)

    temp_file = f"comentarios_{process_id}.xlsx"
    existing_comment_links = cargar_comentarios_existentes(file_path, temp_file)

    book, sheet = crear_o_cargar_archivo_temporal(temp_file)

    df_posts = pd.read_excel(file_path, sheet_name="Posts")
    df_posts_dict = df_posts.set_index("Enlace_Post").to_dict("index")

    return book, df_posts_dict, driver, existing_comment_links, sheet, start_time, temp_file


def procesar_comentario_final(cantidad_comentarios, comment_count, existing_comment_links,
                              first_comment, post_link, sheet, username_autor):
    """
    Procesa un último comentario si no se alcanzó el total y no es del autor.
    """
    if (
            comment_count < cantidad_comentarios
            and first_comment
            and first_comment["author"].lower() != f"@{username_autor.lower()}"
    ):
        guardar_comentario(first_comment, post_link, "No", sheet)
        existing_comment_links.add(first_comment["link"])
        comment_count += 1

    return comment_count


def guardar_comentario(comment, post_link, respuesta, sheet):
    """
    Agrega un comentario formateado a la hoja de Excel.
    """
    sheet.append([
        post_link,
        comment["link"],
        comment["content"],
        comment["date"],
        *comment["metricas"],
        respuesta
    ])


def crear_o_cargar_archivo_temporal(temp_file):
    """
    Crea o carga un archivo temporal para almacenar comentarios durante el proceso.
    """
    if os.path.exists(temp_file):
        book = load_workbook(temp_file)
        sheet = book["Comentarios"]
    else:
        book = Workbook()
        sheet = book.active
        sheet.title = "Comentarios"
        sheet.append([
            "Enlace_Post", "Enlace_Comentario", "Contenido", "Fecha_Publicación",
            "Comentarios", "Retweets", "Likes", "Respuesta"
        ])
        book.save(temp_file)

    return book, sheet


def cargar_comentarios_existentes(file_path, temp_file):
    """
    Carga los enlaces de comentarios ya extraídos (tanto del archivo final como del temporal).
    """
    existing_links = set()
    for ruta in [file_path, temp_file]:
        if os.path.exists(ruta):
            try:
                df = pd.read_excel(ruta, sheet_name="Comentarios")
                existing_links.update(df["Enlace_Comentario"].dropna().tolist())
            except Exception:
                pass
    return existing_links
